# import needed libraries
import time
import os
import re

import numpy as np
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import DesiredCapabilities

from bs4 import BeautifulSoup

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

#================================================================================
#================================================================================
#================================================================================


class color:
    '''
    class to print out colored, bold and underlined text
    '''
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    BLUE = '\033[34m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

#================================================================================
#================================================================================
#================================================================================


def create_dir_options(dir_in):
    if dir_in:
        dir_options_out = ['../'] + \
            sorted([i for i in os.listdir(dir_in)
                    if os.path.isdir(os.path.join(dir_in, i))]) + \
            sorted([i for i in os.listdir(dir_in)
                    if not os.path.isdir(os.path.join(dir_in, i))
                    and 'DS' not in i and '.ipynb' not in i and '.py' not in i])
    else:
        dir_options_out = []

    return(dir_options_out)

#================================================================================
#================================================================================
#================================================================================

def wait_for_downloads(dir_in='/Users/drk36/Downloads', list_in=None):

    print(5 * ' ' + "Waiting for downloads", end="")

    list_new = os.listdir(dir_in)

    if list_in:
        while list_new == list_in:
            time.sleep(2)
            list_new = os.listdir(dir_in)
            print(".", end="")

    while any([filename.endswith(".crdownload") for filename in list_new]):
        time.sleep(2)
        print(".", end="")

    print(" Done!")

#================================================================================
#================================================================================
#================================================================================


def search_wos(all_db_in=None,
               search_string_in=None,
               start_date_in=None,
               end_date_in=None,
               filter_in=None,
               doctype_in=None,
               save_recs_in=None,
               fout_type_in=None,
               wait_click_in=3
               ):

    #------------------------------------------------------------
    # start timing code
    ti = time.time()

    url_in = 'http://apps.webofknowledge.com/'

    wait_click = wait_click_in  # time in seconds to wait for elements to be found by selenium
    stop_option = False  # option for process repeat until successful try
    try_number = 1  # starting try number
    print(f'\n{color.BOLD}{color.BLUE}Starting try: {color.RED}{try_number}{color.END}')
    print(f'\n{color.BOLD}{color.BLUE}output file type: {color.RED}{fout_type_in}{color.END}')

    while not stop_option and try_number <= 3:
        #-----------------------------------
        # start try to find and extract desired records
        try:
            # set options for webdriver
            optionz = webdriver.ChromeOptions()
            optionz.add_argument('--ignore-certificate-errors')
            optionz.add_argument('--incognito')
            capabilities = DesiredCapabilities.CHROME.copy()
            capabilities['acceptSslCerts'] = True
            capabilities['acceptInsecureCerts'] = True

            # start webdriver
            browser = webdriver.Chrome(options=optionz, desired_capabilities=capabilities)
            browser.get(url_in)

            # close popups at the beginning
            close_xpath = '//button[text()="×"]'
            wait = WebDriverWait(browser, 10)
            try:
                wait.until(EC.element_to_be_clickable((By.XPATH, close_xpath)))
                time.sleep(wait_click)
                browser.find_element_by_xpath(close_xpath).click()

                try:
                    wait.until(EC.element_to_be_clickable((By.XPATH, close_xpath)))
                    time.sleep(wait_click)
                    browser.find_element_by_xpath(close_xpath).click()
                except:
                    pass

            except:
                pass

            #-----
            # select option to search in all available databases
            # and not just the default option which is WOS core database
            if all_db_in:
                database_xpath = '//*[@id="global-select"]/div/div[2]/div[1]'
                wait = WebDriverWait(browser, 10)
                wait.until(EC.element_to_be_clickable((By.XPATH, database_xpath)))
                time.sleep(wait_click)
                browser.find_element_by_xpath(database_xpath).click()

            #-----
            # go to advanced search tab
            adv_search_xpath = '//*[@id="snSearchType"]/div[2]/a'
            wait.until(EC.element_to_be_clickable((By.XPATH, adv_search_xpath)))
            time.sleep(wait_click)
            browser.find_element_by_xpath(adv_search_xpath).click()
            time.sleep(2)

            #-----
            # set filter for start year of publications
            if start_date_in:

                add_date_xpath = '//span[text()=" Add date range "]'
                browser.find_element_by_xpath(add_date_xpath).click()
                time.sleep(1)

                try:
                    date_type_xpath = '//span[text()="Index Date"]'
                    browser.find_element_by_xpath(date_type_xpath).click()
                    time.sleep(1)
                except:
                    pass

                date_type_xpath2 = '//span[text()="Publication Date"]'
                browser.find_element_by_xpath(date_type_xpath2).click()
                time.sleep(1)

                try:
                    custom_date_xpath = '//span[text()="Custom"]'
                    browser.find_element_by_xpath(custom_date_xpath).click()
                    time.sleep(1)
                except:
                    pass

                start_date_box_xpath = '/html/body/app-wos/div/div/main/div/app-input-route/' + \
                                       'app-search-home/div[2]/div/app-input-route/app-search-advanced/' + \
                                       'app-advanced-search-form/form/div[4]/div[1]/div[1]/' + \
                                       'app-search-timespan/div/div[2]/input[1]'

                start_date_box = browser.find_element_by_xpath(start_date_box_xpath)
                start_date_box.send_keys(start_date_in)
                time.sleep(1)

                # set filter for end year of publications
                if end_date_in:

                    end_date_box_xpath = '/html/body/app-wos/div/div/main/div/app-input-route/' + \
                                         'app-search-home/div[2]/div/app-input-route/app-search-advanced/' + \
                                         'app-advanced-search-form/form/div[4]/div[1]/div[1]/' + \
                                         'app-search-timespan/div/div[2]/input[2]'

                    end_date_box = browser.find_element_by_xpath(end_date_box_xpath)
                    end_date_box.send_keys(end_date_in)
                    time.sleep(1)

            #-----
            # find and clear search box
            search_xpath = '//*[@id="advancedSearchInputArea"]'
            search_box = browser.find_element_by_xpath(search_xpath)
            search_box.clear()

            # enter search string in search box
            search_box.send_keys(search_string_in)
            search_box.submit()

            try:
                wait.until(EC.element_to_be_clickable((By.XPATH, close_xpath)))
                time.sleep(wait_click)
                browser.find_element_by_xpath(close_xpath).click()

            except:
                pass

            # filter for highly cited publications
            if filter_in != 'All':
                filter_in_xpath = f'//span[text()="{filter_in}"]'
                wait.until(EC.element_to_be_clickable((By.XPATH, filter_in_xpath)))
                browser.find_element_by_xpath(filter_in_xpath).click()

                refine_xpath = '//span[text()="Refine"]'
                wait.until(EC.element_to_be_clickable((By.XPATH, refine_xpath)))
                browser.find_element_by_xpath(refine_xpath).click()

                time.sleep(5)

            # filter for document type
            if doctype_in != 'All':
                doctype_in_xpath = f'//span[text()="{doctype_in}"]'
                wait.until(EC.element_to_be_clickable((By.XPATH, doctype_in_xpath)))
                browser.find_element_by_xpath(doctype_in_xpath).click()

                refine_xpath = '//span[text()="Refine"]'
                wait.until(EC.element_to_be_clickable((By.XPATH, refine_xpath)))
                browser.find_element_by_xpath(refine_xpath).click()

                time.sleep(5)

            # get the number of results after filtering (if any...)
            hits_element_xpath = '/html/body/app-wos/div/div/main/div/app-input-route/app-base-summary-component/' + \
                                 'app-search-friendly-display/div[1]/app-general-search-friendly-display/h1/span'
            hits_element_final = browser.find_element_by_xpath(hits_element_xpath)
            num_results = int(hits_element_final.text.replace(',', ''))

            #------------------------------
            # set list of records' batches to download
            # (only 1000 records to download at a time is allowed)
            #------------------------------
            steps_used = 1000
            if save_recs_in == 'all':
                if num_results % steps_used:
                    recs_used = [i for i in range((num_results // steps_used) + 1)]
                else:
                    recs_used = [i for i in range(num_results // steps_used)]

            elif '-' in save_recs_in:
                save_recs_in = save_recs_in.split('-')
                save_recs_in = [int(i.strip()) if i.strip() else 0 for i in save_recs_in]
                recs_min = save_recs_in[0]
                recs_max = save_recs_in[1]

                while recs_max * steps_used > num_results:
                    recs_max -= 1

                recs_used = [i for i in range(recs_min, recs_max + 1)]

            else:
                recs_used = [0]

            # find number of total pages with results
            total_pages_xpath = '/html/body/app-wos/div/div/main/div/app-input-route/app-base-summary-component/' + \
                                'div/div[2]/app-page-controls[1]/div/form/div/span'
            total_pages_final = browser.find_element_by_xpath(total_pages_xpath)
            num_pages = int(total_pages_final.text.replace(',', ''))

            print(f" - Number of results: {color.BOLD}{color.RED}{num_results}{color.END}, " +
                  f"in {color.BOLD}{color.RED}{num_pages}{color.END} pages")

            # ================================
            # OUTPUT RESULTS IN FILE
            # ================================
            # loop over desired batches of records
            for irec, rec in enumerate(recs_used):
                element_number = irec * 2

                export_continue = True
                while export_continue:
                    markFrom = rec * steps_used + 1
                    if markFrom < num_results:
                        markTo = (rec + 1) * steps_used
                        if markTo > num_results:
                            markTo = num_results

                        print(f' - Exporting records: {markFrom} – {markTo}')

                        export_continue = False
                    else:
                        rec -= 1

                # open up export popup window and make selections
                export_type_xpath = '//span[text()=" Export "]'
                browser.find_element_by_xpath(export_type_xpath).click()
                time.sleep(0.5)

                if fout_type_in == '.xls':
                    # select other format of output file
                    xls_format_xpath = '//button[text()=" Excel "]'
                    browser.find_element_by_xpath(xls_format_xpath).click()
                else:
                    # select other format of output file
                    other_format_xpath = '//button[text()=" Tab delimited file "]'
                    browser.find_element_by_xpath(other_format_xpath).click()

                # select option to save all results
                time.sleep(1)

                save_option_xpath = '//*[@id="radio3"]/label'
                browser.find_element_by_xpath(save_option_xpath).click()

                # input starting record to be exported to file
                markFrom_box = browser.find_element_by_id('mat-input-' + str(element_number))
                markFrom_box.clear()
                markFrom_box.send_keys(markFrom)

                # input ending record to be exported to file
                markTo_box = browser.find_element_by_id('mat-input-' + str(element_number + 1))
                markTo_box.clear()
                markTo_box.send_keys(markTo)

                time.sleep(0.5)
                record_xpath = '//span[text()="Author, Title, Source"]'
                browser.find_element_by_xpath(record_xpath).click()

                time.sleep(0.5)
                record_xpath2 = '//span[text()="Full Record"]'
                browser.find_element_by_xpath(record_xpath2).click()

                time.sleep(0.5)
                export_xpath = '//span[text()="Export"]'
                browser.find_element_by_xpath(export_xpath).click()

                # select option to save all results
                time.sleep(0.5)

                init_list = os.listdir('/Users/drk36/Downloads/')
                wait_for_downloads(list_in=init_list)

            #--------------------
            # end process after successful extraction of records and close webdriver
            stop_option = True
            print(f'{color.BOLD}{color.BLUE}Try: ' +
                  f'{color.RED}{try_number}' +
                  f'{color.BLUE} Successful!!!{color.END}'
                  )
            browser.close()

        #-----------------------------------
        # handles case of falied attempt and repeats process...
        except:
            # close window in  the case of an error
            print(f'{color.BOLD}{color.BLUE}Try ' +
                  f'{color.RED}{try_number} ' +
                  f'{color.BLUE}Failed!!!\n\n' +
                  f'{color.BLUE}Starting try: ' +
                  f'{color.RED}{try_number + 1}{color.END}'
                  )
            try_number += 1

    #--------------------------------------------------
    # end timing code
    tf = time.time()
    print(f'\n{color.BOLD}{color.BLUE}Total time: ' +
          f'{color.RED}{round(tf-ti, 1)} ' +
          f'{color.BLUE}seconds{color.END}\n')

#================================================================================
#================================================================================
#================================================================================


def create_regex_function(data_in,
                          highlight_values_option=False,
                          highlight_text_option=False,
                          count_text_option=False,
                          ):

    #––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
    search_regex = {i: '' for i in data_in.keys()}
    count_words_out = []

    #---------------
    if highlight_values_option:

        if 'values' in data_in.keys():
            df_values = data_in['values'].copy()
            df_values.set_index('name', drop=True, inplace=True)

            #-----
            search_values_numbers = '(-|−)?' + ''.join([str(i) for i in df_values.loc['numbers']
                                                        if str(i) != 'nan'])
            search_values_connect = '|'.join([i for i in df_values.loc['connect'] if str(i) != 'nan'])
            search_values_connect = search_values_connect.replace('+', '\+')
            search_values_prefix = '|'.join([i for i in df_values.loc['prefix'] if str(i) != 'nan'])
            search_values_units = '|'.join([i for i in df_values.loc['units'] if str(i) != 'nan'])
            search_all_units = f'({search_values_prefix})?({search_values_units})' + \
                               f'(\s)?(GAE)?(/)?({search_values_prefix})?({search_values_units})?((-|−)1)?'

            search_values_regex1 = f'({search_values_numbers})(\s)?({search_values_connect})?(\s)?' + \
                                   f'({search_values_numbers})?(\s)?({search_all_units})'
            search_values_regex2 = f'({search_values_numbers})(\s)?({search_values_connect})(\s)?' + \
                                   f'({search_values_numbers})'
            search_values_regex = f'({search_values_regex1}|{search_values_regex2})'

            search_regex['values'] = r'%s' % search_values_regex

    #---------------
    if highlight_text_option:
        #-----
        if 'wanted' in data_in.keys():
            search_wanted_words = [r'%ss?' % (str(i))
                                   for k in data_in['wanted'].columns
                                   for i in data_in['wanted'][k] if str(i) != 'nan'
                                   ]

            search_wanted_regex = '|'.join(search_wanted_words)

            search_wanted_regex = search_wanted_regex.replace('(', '\(').replace(')', '\)')
            search_wanted_regex = search_wanted_regex.replace('[', '\[').replace(']', '\]')
            search_wanted_regex = search_wanted_regex.replace(',', '\,').replace('.', '\.')
            search_wanted_regex = search_wanted_regex.replace('%', '\%').replace(':', '\:')
            search_wanted_regex = search_wanted_regex.replace(' | ', '|')

            search_wanted_regex = search_wanted_regex.replace('\\[-', '[-')
            search_wanted_regex = search_wanted_regex.replace('\\s\\]', '\\s]')

            search_regex['wanted'] = r'%s' % search_wanted_regex

        #-----
        if 'unwanted' in data_in.keys():
            search_unwanted_words = ['|'.join([r'%ss?' % (str(j)) for j in i if str(j) != 'nan'])
                                     for i in data_in['unwanted'].values if str(i[0]) != 'nan']

            search_unwanted_regex = '|'.join(search_unwanted_words)
            search_unwanted_regex = search_unwanted_regex.replace('(', '\(').replace(')', '\)')
            search_unwanted_regex = search_unwanted_regex.replace('[', '\[').replace(']', '\]')
            search_unwanted_regex = search_unwanted_regex.replace(',', '\,').replace('.', '\.')
            search_unwanted_regex = search_unwanted_regex.replace('%', '\%').replace(':', '\:')
            search_unwanted_regex = search_unwanted_regex.replace(' | ', '|')

            search_unwanted_regex = search_unwanted_regex.replace('\\[-', '[-')
            search_unwanted_regex = search_unwanted_regex.replace('\\s\\]', '\\s]')

            search_regex['unwanted'] = r'%s' % search_unwanted_regex

    #---------------
    if count_text_option:
        count_words_out = [str(i) + 's?'
                           for k in data_in['wanted'].columns
                           for i in data_in['wanted'][k] if str(i) != 'nan'
                           ]

    #––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
    return(search_regex, count_words_out)
    #––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––

#================================================================================
#================================================================================
#================================================================================


def read_soup_function(url_in,
                       hide_webdriver_in=True,
                       web_options_in=None,
                       ):

    soup_out = None
    soup_url = None
    success_out = False

    #-------------------------
    if url_in:
        if 'http' in url_in:
            url_used = url_in
        else:
            url_used = f'http://doi.org/{url_in}'

        #-----
        hide_webdriver = hide_webdriver_in

        if web_options_in:
            webdriver_options = web_options_in
        else:
            webdriver_options = webdriver.ChromeOptions()
            if hide_webdriver:
                webdriver_options.add_argument('-headless')
            webdriver_options.add_argument('--ignore-certificate-errors')
            webdriver_options.add_argument('--incognito')

        #--------------------------------------------------
        # Get soup of paper from the publisher's website
        #--------------------------------------------------

        print(color.BOLD, color.BLUE, 100 * '-', color.END)
        print(f'{color.BOLD}{color.BLUE}  Starting try to read paper ' +
              f'{url_in}!{color.END}\n')

        #---------------
        driver = webdriver.Chrome(options=webdriver_options)
        driver.get(url_used)
        time.sleep(3)

        #---------------
        accept_xpath = '//a[text()="Accept"]'
        if driver.find_elements_by_xpath(accept_xpath):
            try:
                wait = WebDriverWait(driver, 10)
                wait.until(EC.element_to_be_clickable((By.XPATH, accept_xpath)))
                driver.find_element_by_xpath(accept_xpath).click()
            except:
                pass

        #---------------
        accept2_xpath = '//button[text()="Accept Cookies"]'
        if driver.find_elements_by_xpath(accept2_xpath):
            try:
                wait = WebDriverWait(driver, 10)
                wait.until(EC.element_to_be_clickable((By.XPATH, accept2_xpath)))
                driver.find_element_by_xpath(accept2_xpath).click()
            except:
                pass

        #---------------
        allow_xpath = '//a[class="optanon-allow-all"]'
        if driver.find_elements_by_xpath(allow_xpath):
            wait = WebDriverWait(driver, 10)
            wait.until(EC.element_to_be_clickable((By.XPATH, allow_xpath)))
            driver.find_element_by_xpath(allow_xpath).click()

        #---------------
        view_xpath = '//a[text()="View Full-Text"]'
        if driver.find_elements_by_xpath(view_xpath):
            wait = WebDriverWait(driver, 10)
            wait.until(EC.element_to_be_clickable((By.XPATH, view_xpath)))
            driver.find_element_by_xpath(view_xpath).click()

        time.sleep(2)

        #-------------------------
        soup_url = driver.current_url
        print(f'\t{color.BOLD}{color.RED}URL: {color.END}', soup_url, '\n')

        if '.pdf' in soup_url:
            comment_out += 'pdf_file | '

        else:

            #---------------
            if 'oeno-one' in soup_url:
                expand_css = 'a[class*="tab-expand"]'
                tab_ex_els = driver.find_elements_by_css_selector(expand_css)
                if tab_ex_els:
                    for itabex in range(len(tab_ex_els)):
                        tab_ex_els[itabex].click()
                        time.sleep(0.5)

            #---------------
            soup_html = driver.execute_script('return document.documentElement.outerHTML')
            soup_out = BeautifulSoup(soup_html, "html.parser")

            success_out = True

        #-------------------------
        driver.close()

    #-------------------------
    return(soup_url, soup_out, success_out)

#================================================================================
#================================================================================
#================================================================================

def add_counts_function(df_in):

    if not df_in.empty:

        df_out = df_in.copy()

        if 'col_count' in df_out.index:
            start_row = 1
        else:
            start_row = 0

        if all([i.isnumeric() for j in df_out.iloc[start_row:, 2:].replace('', '0').values for i in j]):
            df_type = 'numeric'
        else:
            df_type = 'text'

        if 'row_count' not in df_out.columns:
            if df_type == 'numeric':
                df_out.iloc[start_row:, 2:] = df_out.iloc[start_row:, 2:].replace('', 0).astype('int32')
                df_out['row_count'] = df_out.iloc[start_row:, 2:].astype(bool).sum(axis=1)
            elif df_type == 'text':
                df_out['row_count'] = \
                    [len([j for j in df_out.iloc[i, 2:].values if j != '']) for i in range(df_out.shape[0])]

        df_out = df_out[df_out['row_count'].values > 0.]

        if df_type == 'numeric':
            count_list = list(df_out.iloc[start_row:, 2:].astype(bool)
                                                         .sum(axis=0))

        elif df_type == 'text':
            count_list = \
                [len([j for j in df_out.iloc[start_row:, i].values if j != ''])
                 for i in range(2, df_out.shape[1] - 1)] + \
                [df_out.iloc[start_row:, -1].astype('int32').astype(bool).sum()]

        df_out.loc['col_count'] = ['', ''] + count_list
        df_out = df_out.reindex(index=['col_count'] + [i for i in df_out.index if i != 'col_count'])

        return(df_out)

#================================================================================
#================================================================================
#================================================================================

def df_to_excel_function_new(io_in, sheetname_in, df_in, emptyrows_in = 2):
    """
    Check if sheet exists for an xlsx spreadsheet and add data from dataframe to the sheet
    :param: io_in - The io_in of the xlsx spreadsheet
    :param: sheetname_in - Name of the worksheet to search for
    :param: df_in - A Pandas dataframe object
    
    """
    #----------
    if not os.path.exists(io_in):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheetname_in
    #-----
    else:
        wb = openpyxl.load_workbook(io_in)
    
    
    #----------
    try:
        ws = wb[sheetname_in]
    
    #-----
    except KeyError:
        wb.create_sheet(sheetname_in)
        ws = wb[sheetname_in]
    
    #-----
    finally:
        maxrow = wb[sheetname_in].max_row
        
        # append empty rows to excel sheet if sheet not empty
        if maxrow > 1:
            df_empty = pd.DataFrame(index = [f'temp{i}' for i in range(emptyrows_in)], 
                                    columns = df_in.columns)
            for r in dataframe_to_rows(df_empty, index = False, header = False):
                ws.append(r)
                
            maxrow += emptyrows_in + 1
        
        # append dataframe rows
        for ir, r in enumerate(dataframe_to_rows(df_in, index = True, header = True)):
            if ir != 1:
                ws.append(r)

        ws.cell(column = 1, row = maxrow, value = df_in.index.name)
            
        # change "index" and "columns" styling
        for cell in ws['A'][-df_in.shape[0]:] + ws[maxrow]:
            cell.style = 'Pandas'
        
        wb.save(io_in)

#================================================================================
#================================================================================
#================================================================================
